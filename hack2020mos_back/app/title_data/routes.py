from openpyxl import load_workbook
from flask import request
from flask_restx import Resource, Namespace, fields, reqparse
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from app.title_data.utils import allowed_file, ALLOWED_EXTENSIONS
from extensions import db
from models import Road as RoadModel, TitleDataSchema
from models import TitleData as TitleDataModel

ns = Namespace('TitleData', description='Загрузка титульного листа', path='/title-data')

model_output = ns.model('TitleDataOutput', {
    'id': fields.Integer,
    'name': fields.String,
    'upload_date': fields.DateTime,
    'objects_count': fields.Integer
})

model_list_output = ns.model('TitleDataListOutput', {
    'pageItems': fields.List(fields.Nested(model_output)),
    'total': fields.Integer
})

model_output_upload = ns.model('TitleDataUploadOutput', {
    'address_count': fields.Integer,
    'address_count_unique': fields.Integer
})

upload_parser = ns.parser()
upload_parser.add_argument('file', type=FileStorage, location='files', required=True)

search_parser = reqparse.RequestParser()
search_parser.add_argument('name', location='json')
search_parser.add_argument('upload_date', type=str, location='json')


@ns.route('/search')
class TitleDataSearch(Resource):

    @ns.doc(responses={500: 'Неизвестная ошибка'},
            params={'per_page': 'Кол-во записей на странице', 'page': 'Номер страницы'},
            body=search_parser)
    @ns.response(200, 'OK', model_list_output)
    def post(self):
        """ Получить загруженные титульные списки """
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per-page', 10, type=int)

        search_data = search_parser.parse_args()
        search_data = {k: v for k, v in search_data.items() if v not in [[], '', None]}

        rows_count = db.session.query(TitleDataModel).count()
        reports = []

        if search_data == {}:
            reports = TitleDataModel.query.order_by(TitleDataModel.created_date.desc()).paginate(page, per_page, False)\
                .items

        return {
            'page_items': TitleDataSchema().dump(reports, many=True),
            'total': rows_count
        }


@ns.route('')
class TitleData(Resource):

    @ns.doc(responses={500: 'Неизвестная ошибка'},
            body=upload_parser)
    @ns.response(200, 'OK', model_output_upload)
    def post(self):
        """
        Загрузка титульного листа
        """

        args = upload_parser.parse_args()
        uploaded_file = args['file']

        if not allowed_file(uploaded_file.filename):
            return {
                "message": "Можно загрузить файлы только в форматах {}".format(', '.join(ALLOWED_EXTENSIONS))
            }

        if uploaded_file:
            filename = secure_filename(uploaded_file.filename)
            wb = load_workbook(uploaded_file, read_only=True)
            sheet = wb.worksheets[0]
            rows = sheet.max_row
            rows_unique = 0

            title_data = TitleDataModel(
                name=filename,
                objects_count=rows - 1
            )
            db.session.add(title_data)

            for i in range(7, rows + 1):
                address = sheet.cell(row=i, column=2).value
                address_from_bd = RoadModel.query.filter_by(name=address).first()
                if not address_from_bd:
                    rows_unique += 1
                    road = RoadModel(name=address)
                    db.session.add(road)

            db.session.commit()

            return {
                "address_count": rows - 1,
                "address_count_unique": rows_unique
            }
